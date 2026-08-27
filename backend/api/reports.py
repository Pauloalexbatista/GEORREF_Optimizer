from fastapi import APIRouter, HTTPException, Depends, status
from fastapi.responses import StreamingResponse
from typing import Dict, Any, List
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from database import get_db, get_projeto
from backend.api.auth import get_current_user, UserResponse
from utils.persistence_manager import serialize_state, deserialize_state

router = APIRouter(prefix="/reports", tags=["reports"])

@router.get("/{project_id}/summary")
def get_project_reports_summary(project_id: int, current_user: UserResponse = Depends(get_current_user)):
    proj = get_projeto(project_id)
    if not proj:
        raise HTTPException(status_code=404, detail="Projeto não encontrado.")
    if proj["empresa_id"] != current_user.empresa_id and not getattr(current_user, "is_superadmin", False):
        raise HTTPException(status_code=403, detail="Sem permissão para aceder a este projeto.")
        
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT payload_json FROM snapshots WHERE projeto_id = ? ORDER BY id DESC LIMIT 1", (project_id,))
        row = cursor.fetchone()
        
        if not row:
            return {"empty": True}
            
        state_dict = deserialize_state(row["payload_json"])
        raw_routes = state_dict.get("routes_solution")
        
        if raw_routes is None:
            return {"empty": True}
            
        df_routes = raw_routes if isinstance(raw_routes, pd.DataFrame) else pd.DataFrame(raw_routes)
        if df_routes.empty:
            return {"empty": True}
            
        total_stops = len(df_routes)
        entregues = int((df_routes.get("Estado", pd.Series(["Pendente"]*total_stops)) == "Entregue").sum())
        falhadas = int((df_routes.get("Estado", pd.Series(["Pendente"]*total_stops)) == "Não Entregue").sum())
        pendentes = total_stops - entregues - falhadas
        rate = round((entregues / total_stops * 100)) if total_stops > 0 else 0
        
        total_km = round(float(df_routes["KM_Anterior"].fillna(0).astype(float).sum()), 1) if "KM_Anterior" in df_routes.columns else 0.0
        total_weight = round(float(df_routes["Peso"].fillna(0).astype(float).sum()), 1) if "Peso" in df_routes.columns else 0.0
        total_packages = int(df_routes["Volumes"].fillna(1).astype(int).sum()) if "Volumes" in df_routes.columns else total_stops
        
        # Route breakdown
        route_stats = []
        for r_name, group in df_routes.groupby("Rota"):
            if str(r_name).lower() in ["por distribuir", "pendente", "nan", ""]:
                continue
            r_total = len(group)
            r_ent = int((group.get("Estado", pd.Series(["Pendente"]*r_total)) == "Entregue").sum())
            r_fal = int((group.get("Estado", pd.Series(["Pendente"]*r_total)) == "Não Entregue").sum())
            r_km = round(float(group["KM_Anterior"].fillna(0).astype(float).sum()), 1) if "KM_Anterior" in group.columns else 0.0
            r_weight = round(float(group["Peso"].fillna(0).astype(float).sum()), 1) if "Peso" in group.columns else 0.0
            
            route_stats.append({
                "route_name": str(r_name),
                "total": r_total,
                "entregues": r_ent,
                "falhadas": r_fal,
                "pendentes": r_total - r_ent - r_fal,
                "rate": round((r_ent / r_total * 100)) if r_total > 0 else 0,
                "km": r_km,
                "weight": r_weight
            })
            
        # Failure reasons count
        reasons_dist = {}
        if "Motivo_Falha" in df_routes.columns:
            val_counts = df_routes[df_routes["Motivo_Falha"] != ""]["Motivo_Falha"].value_counts().to_dict()
            reasons_dist = {str(k): int(v) for k, v in val_counts.items()}
            
        return {
            "empty": False,
            "project_name": proj["nome"],
            "totals": {
                "total_stops": total_stops,
                "entregues": entregues,
                "falhadas": falhadas,
                "pendentes": pendentes,
                "rate": rate,
                "total_km": total_km,
                "total_weight": total_weight,
                "total_packages": total_packages,
                "total_routes": len(route_stats)
            },
            "routes": route_stats,
            "reasons": reasons_dist
        }

@router.get("/{project_id}/export")
def export_project_final_report(project_id: int, current_user: UserResponse = Depends(get_current_user)):
    proj = get_projeto(project_id)
    if not proj:
        raise HTTPException(status_code=404, detail="Projeto não encontrado.")
        
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT payload_json FROM snapshots WHERE projeto_id = ? ORDER BY id DESC LIMIT 1", (project_id,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=400, detail="Sem dados para exportar.")
            
        state_dict = deserialize_state(row["payload_json"])
        raw_routes = state_dict.get("routes_solution")
        if raw_routes is None:
            raise HTTPException(status_code=400, detail="Sem rotas para exportar.")
            
        df_routes = raw_routes if isinstance(raw_routes, pd.DataFrame) else pd.DataFrame(raw_routes)
        
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws_rep = wb.active
    ws_rep.title = "Relatório de distribuição"
    
    headers = [
        "Rota", "Ordem", "Cliente", "Morada", "Localidade", "CodPostal", 
        "Contacto", "Janela Prevista", "Hora Picagem", "Estado", 
        "Motivo de Não Entrega", "Notas do Motorista", "Peso (kg)", "Volumes"
    ]
    ws_rep.append(headers)
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    
    for col_num in range(1, len(headers) + 1):
        cell = ws_rep.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for _, r in df_routes.iterrows():
        ws_rep.append([
            str(r.get("Rota", "")),
            int(r.get("Ordem", 0)) if pd.notna(r.get("Ordem", 0)) else "",
            str(r.get("Cliente", "")),
            str(r.get("Morada", "")),
            str(r.get("Localidade", "")),
            str(r.get("CodPostal", r.get("Cod_Postal", ""))),
            str(r.get("Contacto", r.get("Telefone", ""))),
            f"{r.get('Janela_Inicio', '08:00')} - {r.get('Janela_Fim', '18:00')}",
            str(r.get("Hora_Picagem", "")),
            str(r.get("Estado", "Pendente")),
            str(r.get("Motivo_Falha", "")),
            str(r.get("Notas_Motorista", "")),
            float(r.get("Peso", 0.0)) if pd.notna(r.get("Peso", 0.0)) else 0.0,
            int(r.get("Volumes", 1)) if pd.notna(r.get("Volumes", 1)) else 1
        ])
        
    wb.save(output)
    output.seek(0)
    
    filename = f"Relatorio_Final_{proj['nome'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
