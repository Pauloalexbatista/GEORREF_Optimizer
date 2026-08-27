import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
from typing import Dict, Any, List
from .models import get_db, set_session_meta, get_session_meta, clear_session_db

# Standard GeoRoutePlan Color Palette
FONT_FAMILY = "Segoe UI"
FONT_HEADER = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
FONT_DATA = Font(name=FONT_FAMILY, size=10, color="1F2937")

COLOR_REQ = "1E3A8A"       # Azul Escuro (OBRIGATÓRIO)
COLOR_REC = "2563EB"       # Azul Mais Claro (RECOMENDADO)
COLOR_OPT = "64748B"       # Cinza Ardósia (OPCIONAL)
COLOR_INSTRUCOES = "1E293B"# Ardósia Escuro
COLOR_SUMMARY = "0F766E"   # Verde Petróleo para Relatório Final

FILL_ZEBRA = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER_SIDE = Side(border_style="thin", color="E2E8F0")
BORDER_CELL = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
BORDER_HEADER = Border(
    left=Side(border_style="thin", color="CBD5E1"),
    right=Side(border_style="thin", color="CBD5E1"),
    top=Side(border_style="medium", color="0F172A"),
    bottom=Side(border_style="medium", color="0F172A")
)

DEFAULT_FAILURE_REASONS = [
    "Cliente Ausente / Fechado",
    "Carga Não Conforme / Danificada",
    "Cliente Recusou a Carga",
    "Morada Incorreta ou Incompleta",
    "Sem Acesso / Rua em Obras",
    "Fora do Horário de Descarga",
    "Falta de Pagamento (Cobrança)",
    "Avaria / Atraso Operacional",
    "Outro Motivo (ver notas)"
]

def clean_str(val):
    if pd.isna(val) or val is None:
        return ""
    return str(val).strip()

def clean_float(val, default=0.0):
    try:
        if pd.isna(val) or val is None or str(val).strip() == "":
            return default
        return float(val)
    except:
        return default

def clean_int(val, default=0):
    try:
        if pd.isna(val) or val is None or str(val).strip() == "":
            return default
        return int(float(val))
    except:
        return default

def parse_and_import_excel(file_path: str, original_filename: str) -> Dict[str, Any]:
    clear_session_db()
    
    with pd.ExcelFile(file_path) as excel_file:
        sheet_names = excel_file.sheet_names
        
        # 1. Sheet: Armazém (se existir)
        sheet_wh_name = None
        for s in sheet_names:
            if "armaz" in s.lower() or "depot" in s.lower() or "polo" in s.lower():
                sheet_wh_name = s
                break
        df_wh = pd.read_excel(excel_file, sheet_name=sheet_wh_name) if sheet_wh_name else None

        # 2. Sheet: Rotas
        sheet_rotas_name = None
        for s in sheet_names:
            if "rota" in s.lower() or "viagen" in s.lower() or "distribui" in s.lower() or "entregas" in s.lower():
                sheet_rotas_name = s
                break
        if not sheet_rotas_name:
            # Fallback to 2nd sheet if Armazém is 1st, or 1st sheet
            sheet_rotas_name = sheet_names[1] if len(sheet_names) > 1 and sheet_wh_name == sheet_names[0] else sheet_names[0]
            
        df_rotas = pd.read_excel(excel_file, sheet_name=sheet_rotas_name)
        
        # 3. Sheet: Motoristas e Carros
        sheet_drivers_name = None
        for s in sheet_names:
            if "motorista" in s.lower() or "condutor" in s.lower() or "driver" in s.lower() or "carros" in s.lower() or "frota" in s.lower():
                sheet_drivers_name = s
                break
        df_drivers = pd.read_excel(excel_file, sheet_name=sheet_drivers_name) if sheet_drivers_name else None
        
        # 4. Sheet: Justificações / Motivos
        sheet_reasons_name = None
        for s in sheet_names:
            if "justifica" in s.lower() or "motivo" in s.lower() or "falha" in s.lower() or "reasons" in s.lower():
                sheet_reasons_name = s
                break
        df_reasons = pd.read_excel(excel_file, sheet_name=sheet_reasons_name) if sheet_reasons_name else None

    # Handle Warehouse metadata
    wh_info_str = "Armazém Principal"
    if df_wh is not None and len(df_wh) > 0:
        df_wh.columns = [str(c).strip() for c in df_wh.columns]
        first_row = df_wh.iloc[0]
        wh_name = clean_str(first_row.get("Nome_Armazem") or first_row.get("Armazem") or first_row.iloc[0])
        wh_addr = clean_str(first_row.get("Morada") or "")
        wh_cp = clean_str(first_row.get("CP") or first_row.get("Codigo_Postal") or "")
        wh_info_str = f"{wh_name} ({wh_addr}, {wh_cp})".strip()
        set_session_meta("warehouse_name", wh_name)
        set_session_meta("warehouse_address", wh_addr)
        set_session_meta("warehouse_cp", wh_cp)

    df_rotas.columns = [str(c).strip() for c in df_rotas.columns]
    
    conn = get_db()
    cursor = conn.cursor()
    
    def find_col(possible_names):
        for col in df_rotas.columns:
            for name in possible_names:
                if name.lower() in col.lower():
                    return col
        return None

    col_route = find_col(["rota", "route", "id_rota", "viagem", "carro", "veiculo"]) or df_rotas.columns[0]
    col_seq = find_col(["seq", "ordem", "paragem", "stop", "pos"])
    col_client = find_col(["cliente", "nome", "destinatario", "empresa"]) or df_rotas.columns[min(1, len(df_rotas.columns)-1)]
    col_address = find_col(["morada", "endereco", "rua", "address"])
    col_postal = find_col(["postal", "cp", "cod_postal", "zip"])
    col_city = find_col(["localidade", "cidade", "concelho", "freguesia", "city"])
    col_phone = find_col(["telefone", "tlm", "telemovel", "contacto", "phone"])
    col_contact = find_col(["contacto_nome", "responsavel", "pessoa"])
    col_vol = find_col(["vol", "volume", "m3"])
    col_weight = find_col(["peso", "kg", "weight"])
    col_packages = find_col(["bulto", "caixa", "volume_qtd", "packages", "qtd", "volumes"])
    col_seller = find_col(["vendedor", "comercial", "seller"])
    col_notes = find_col(["obs", "observa", "nota", "notas", "comentario"])
    col_cod = find_col(["cobranca", "cobrar", "valor", "preco", "cod", "pagamento"])

    total_stops = 0
    unique_routes = set()

    for idx, row in df_rotas.iterrows():
        route_id = clean_str(row.get(col_route))
        client = clean_str(row.get(col_client))
        if not route_id and not client:
            continue
        if not route_id:
            route_id = "Rota 1"
            
        unique_routes.add(route_id)
        total_stops += 1
        
        seq = clean_int(row.get(col_seq), default=idx+1) if col_seq else idx+1
        address = clean_str(row.get(col_address)) if col_address else ""
        postal = clean_str(row.get(col_postal)) if col_postal else ""
        city = clean_str(row.get(col_city)) if col_city else ""
        phone = clean_str(row.get(col_phone)) if col_phone else ""
        contact = clean_str(row.get(col_contact)) if col_contact else ""
        vol = clean_float(row.get(col_vol)) if col_vol else 0.0
        weight = clean_float(row.get(col_weight)) if col_weight else 0.0
        packages = clean_int(row.get(col_packages)) if col_packages else 0
        seller = clean_str(row.get(col_seller)) if col_seller else ""
        notes = clean_str(row.get(col_notes)) if col_notes else ""
        cod = clean_float(row.get(col_cod)) if col_cod else 0.0
        
        cursor.execute("""
            INSERT INTO route_stops (
                route_id, sequence, client_name, address, postal_code, city,
                phone, contact_person, volume, weight, packages, seller,
                notes, cod_amount, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pendente')
        """, (route_id, seq, client, address, postal, city, phone, contact, vol, weight, packages, seller, notes, cod))

    drivers_count = 0
    if df_drivers is not None:
        df_drivers.columns = [str(c).strip() for c in df_drivers.columns]
        
        d_name_col = None
        d_veh_col = None
        d_pass_col = None
        d_route_col = None
        
        for c in df_drivers.columns:
            cl = c.lower()
            if "nome" in cl or "motorista" in cl or "condutor" in cl:
                d_name_col = c
            elif "carro" in cl or "veiculo" in cl or "viatura" in cl or "matricula" in cl:
                d_veh_col = c
            elif "pass" in cl or "pin" in cl or "senha" in cl or "codigo" in cl:
                d_pass_col = c
            elif "rota" in cl:
                d_route_col = c
                
        for idx, row in df_drivers.iterrows():
            name = clean_str(row.get(d_name_col)) if d_name_col else f"Motorista {idx+1}"
            if not name:
                continue
            veh = clean_str(row.get(d_veh_col)) if d_veh_col else ""
            pwd = clean_str(row.get(d_pass_col)) if d_pass_col else f"123{idx+1}"
            assigned = clean_str(row.get(d_route_col)) if d_route_col else ""
            
            cursor.execute("""
                INSERT INTO drivers (name, vehicle, password, assigned_route_id)
                VALUES (?, ?, ?, ?)
            """, (name, veh, pwd, assigned if assigned else None))
            drivers_count += 1
    else:
        for idx, r_id in enumerate(sorted(list(unique_routes))):
            cursor.execute("""
                INSERT INTO drivers (name, vehicle, password, assigned_route_id)
                VALUES (?, ?, ?, ?)
            """, (f"Motorista {idx+1}", f"Viatura {idx+1}", f"{1000 + idx + 1}", r_id))
            drivers_count += 1

    reasons_list = []
    if df_reasons is not None:
        first_col = df_reasons.columns[0]
        for val in df_reasons[first_col].dropna():
            s_val = clean_str(val)
            if s_val and s_val not in reasons_list:
                reasons_list.append(s_val)
                
    if not reasons_list:
        reasons_list = DEFAULT_FAILURE_REASONS

    for r in reasons_list:
        cursor.execute("INSERT INTO failure_reasons (reason) VALUES (?)", (r,))

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.commit()
    conn.close()
    
    set_session_meta("imported_at", now_str)
    set_session_meta("file_name", original_filename)
    set_session_meta("total_stops", str(total_stops))
    set_session_meta("total_routes", str(len(unique_routes)))
    set_session_meta("warehouse_info", wh_info_str)
    set_session_meta("session_active", "1")
    
    return {
        "imported_at": now_str,
        "filename": original_filename,
        "warehouse": wh_info_str,
        "total_stops": total_stops,
        "total_routes": len(unique_routes),
        "total_drivers": drivers_count
    }

def generate_export_excel(output_path: str) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    conn = get_db()
    
    # -------------------------------------------------------------
    # 1. Folha Armazém
    # -------------------------------------------------------------
    ws_wh = wb.create_sheet(title="Armazém")
    wh_name = get_session_meta("warehouse_name") or "Armazém Principal"
    wh_addr = get_session_meta("warehouse_address") or "-"
    wh_cp = get_session_meta("warehouse_cp") or "-"
    
    wh_headers = [
        ("Nome_Armazem", COLOR_REQ),
        ("Morada", COLOR_REQ),
        ("CP", COLOR_REQ),
        ("Status_Operacional", COLOR_REC)
    ]
    wh_data = [
        [wh_name, wh_addr, wh_cp, "Base de Operações Diária"]
    ]
    
    _apply_styled_sheet(ws_wh, wh_headers, wh_data)

    # -------------------------------------------------------------
    # 2. Folha Rotas
    # -------------------------------------------------------------
    ws_rotas = wb.create_sheet(title="Rotas")
    rotas_headers = [
        ("Rota", COLOR_REQ),
        ("Sequência", COLOR_REQ),
        ("Cliente", COLOR_REQ),
        ("Morada", COLOR_REQ),
        ("Código Postal", COLOR_REQ),
        ("Localidade", COLOR_REC),
        ("Contacto", COLOR_REC),
        ("Volume (m3)", COLOR_REC),
        ("Peso (kg)", COLOR_REC),
        ("Bultos/Caixas", COLOR_REC),
        ("Vendedor", COLOR_OPT),
        ("Observações Iniciais", COLOR_OPT),
        ("Valor a Cobrar", COLOR_OPT),
        ("Estado Final", COLOR_REQ),
        ("Motivo Não Entrega", COLOR_REC),
        ("Notas do Motorista", COLOR_REC),
        ("Data/Hora Registo", COLOR_OPT),
        ("Latitude Registo", COLOR_OPT),
        ("Longitude Registo", COLOR_OPT)
    ]
    
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            route_id, sequence, client_name, address, postal_code, city,
            phone, volume, weight, packages, seller, notes, cod_amount,
            status, fail_reason, driver_notes, updated_at, delivered_lat, delivered_lng
        FROM route_stops
        ORDER BY route_id, sequence
    """)
    rotas_data = [list(row) for row in cursor.fetchall()]
    _apply_styled_sheet(ws_rotas, rotas_headers, rotas_data)

    # -------------------------------------------------------------
    # 3. Folha Motoristas e Carros
    # -------------------------------------------------------------
    ws_drivers = wb.create_sheet(title="Motoristas e Carros")
    drivers_headers = [
        ("Motorista", COLOR_REQ),
        ("Viatura", COLOR_REC),
        ("Rota Atribuída", COLOR_REC),
        ("PIN/Password", COLOR_REQ),
        ("Último Sinal GPS", COLOR_OPT),
        ("Última Latitude", COLOR_OPT),
        ("Última Longitude", COLOR_OPT)
    ]
    cursor.execute("""
        SELECT name, vehicle, assigned_route_id, password, last_gps_time, last_lat, last_lng
        FROM drivers
    """)
    drivers_data = [list(row) for row in cursor.fetchall()]
    _apply_styled_sheet(ws_drivers, drivers_headers, drivers_data)

    # -------------------------------------------------------------
    # 4. Folha Justificação entregas
    # -------------------------------------------------------------
    ws_reasons = wb.create_sheet(title="Justificação entregas")
    reasons_headers = [
        ("Motivos de Insucesso", COLOR_REQ)
    ]
    cursor.execute("SELECT reason FROM failure_reasons")
    reasons_data = [[row[0]] for row in cursor.fetchall()]
    _apply_styled_sheet(ws_reasons, reasons_headers, reasons_data)

    # -------------------------------------------------------------
    # 5. Folha Relatório de distribuição
    # -------------------------------------------------------------
    ws_rep = wb.create_sheet(title="Relatório de distribuição")
    rep_headers = [
        ("Rota", COLOR_SUMMARY),
        ("Motorista", COLOR_SUMMARY),
        ("Viatura", COLOR_SUMMARY),
        ("Total Clientes", COLOR_SUMMARY),
        ("Entregues", COLOR_SUMMARY),
        ("Falhadas", COLOR_SUMMARY),
        ("Pendentes", COLOR_SUMMARY),
        ("% Sucesso", COLOR_SUMMARY),
        ("% Falhas", COLOR_SUMMARY),
        ("Primeiro Registo", COLOR_SUMMARY),
        ("Último Registo", COLOR_SUMMARY),
        ("Valor Total Cobrado", COLOR_SUMMARY)
    ]

    cursor.execute("""
        SELECT 
            rs.route_id,
            COUNT(rs.id) as total,
            SUM(CASE WHEN rs.status = 'Entregue' THEN 1 ELSE 0 END) as entregues,
            SUM(CASE WHEN rs.status = 'Não Entregue' THEN 1 ELSE 0 END) as falhadas,
            SUM(CASE WHEN rs.status = 'Pendente' THEN 1 ELSE 0 END) as pendentes,
            MIN(rs.updated_at) as primeira_acao,
            MAX(rs.updated_at) as ultima_acao,
            SUM(CASE WHEN rs.status = 'Entregue' THEN rs.cod_amount ELSE 0 END) as cobrado_total
        FROM route_stops rs
        GROUP BY rs.route_id
        ORDER BY rs.route_id
    """)
    rows = cursor.fetchall()
    
    cursor.execute("SELECT assigned_route_id, name, vehicle FROM drivers WHERE assigned_route_id IS NOT NULL")
    driver_map = {r["assigned_route_id"]: (r["name"], r["vehicle"]) for r in cursor.fetchall()}

    summary_rows = []
    total_all, entregues_all, falhadas_all, pendentes_all, cobrado_all = 0, 0, 0, 0, 0.0

    for r in rows:
        t = r["total"] or 0
        e = r["entregues"] or 0
        f = r["falhadas"] or 0
        p = r["pendentes"] or 0
        cob = r["cobrado_total"] or 0.0
        pct_e = f"{round((e / t * 100), 1)}%" if t > 0 else "0.0%"
        pct_f = f"{round((f / t * 100), 1)}%" if t > 0 else "0.0%"
        drv_name, drv_veh = driver_map.get(r["route_id"], ("Não Atribuído", "-"))
        
        summary_rows.append([
            r["route_id"], drv_name, drv_veh, t, e, f, p, pct_e, pct_f,
            r["primeira_acao"] or "-", r["ultima_acao"] or "-", f"{cob:.2f} €"
        ])
        total_all += t
        entregues_all += e
        falhadas_all += f
        pendentes_all += p
        cobrado_all += cob

    if total_all > 0:
        pct_all_e = f"{round((entregues_all / total_all * 100), 1)}%"
        pct_all_f = f"{round((falhadas_all / total_all * 100), 1)}%"
        summary_rows.append([
            "--- TOTAL GLOBAL ---", "-", "-", total_all, entregues_all, falhadas_all, pendentes_all,
            pct_all_e, pct_all_f, "-", "-", f"{cobrado_all:.2f} €"
        ])

    _apply_styled_sheet(ws_rep, rep_headers, summary_rows)

    conn.close()
    
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path

def _apply_styled_sheet(ws, headers_spec, data_rows):
    ws.views.sheetView[0].showGridLines = True
    
    ws.row_dimensions[1].height = 28
    for col_idx, (col_name, color_hex) in enumerate(headers_spec, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = FONT_HEADER
        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_HEADER

    for row_idx, row_data in enumerate(data_rows, start=2):
        ws.row_dimensions[row_idx].height = 22
        fill = FILL_ZEBRA if row_idx % 2 == 0 else FILL_WHITE
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_DATA
            cell.fill = fill
            cell.border = BORDER_CELL
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
