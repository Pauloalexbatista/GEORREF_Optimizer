import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Universal Color Palette (Standard GeoRoutePlan)
FONT_FAMILY = "Segoe UI"
FONT_HEADER = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
FONT_DATA = Font(name=FONT_FAMILY, size=10, color="1F2937")
FONT_DATA_BOLD = Font(name=FONT_FAMILY, size=10, bold=True, color="1F2937")
FONT_MUTED = Font(name=FONT_FAMILY, size=9, italic=True, color="6B7280")

COLOR_REQ = "1E3A8A"       # Azul Escuro (OBRIGATÓRIO)
COLOR_REC = "2563EB"       # Azul Mais Claro (RECOMENDADO)
COLOR_OPT = "64748B"       # Cinza Ardósia (OPCIONAL)
COLOR_INSTRUCOES = "1E293B"# Ardósia Escuro

FILL_ZEBRA = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER_SIDE = Side(border_style="thin", color="E2E8F0")
BORDER_CELL = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
BORDER_HEADER = Border(
    left=Side(border_style="thin", color="CBD5E1"),
    right=Side(border_style="thin", color="CBD5E1"),
    top=Side(border_style="medium", color="0F172A"),
    bottom=Side(border_style="medium", color="0F172A")
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

def apply_headers_and_data(ws, headers_spec, data_rows):
    ws.views.sheetView[0].showGridLines = True
    
    # 1. Header Row
    ws.row_dimensions[1].height = 28
    for col_idx, (col_name, color_hex) in enumerate(headers_spec, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = FONT_HEADER
        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    # 2. Data Rows
    for row_idx, row_data in enumerate(data_rows, start=2):
        ws.row_dimensions[row_idx].height = 22
        fill = FILL_ZEBRA if row_idx % 2 == 0 else FILL_WHITE
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_DATA
            cell.fill = fill
            cell.border = BORDER_CELL
            if isinstance(val, (int, float)):
                cell.alignment = ALIGN_RIGHT
            else:
                cell.alignment = ALIGN_LEFT

    # 3. Auto column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

def build_app_georouteplan_template(output_path: str):
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # -------------------------------------------------------------
    # 1. FOLHA: Armazém (O Armazém é o 1º!)
    # -------------------------------------------------------------
    ws_wh = wb.create_sheet(title="Armazém")
    wh_headers = [
        ("Nome_Armazem", COLOR_REQ),
        ("Morada", COLOR_REQ),
        ("CP", COLOR_REQ),
        ("Localidade", COLOR_REC),
        ("Contacto", COLOR_REC),
        ("Hora_Abertura", COLOR_REC),
        ("Hora_Fecho", COLOR_REC),
        ("Latitude", COLOR_OPT),
        ("Longitude", COLOR_OPT),
        ("Observacoes", COLOR_OPT)
    ]
    wh_data = [
        ["Armazém Central Porto", "Via do Corvo 150", "4405-555", "Vila Nova de Gaia", "220000111", "07:30:00", "19:30:00", 41.1150, -8.6250, "Cais de carga principal"]
    ]
    apply_headers_and_data(ws_wh, wh_headers, wh_data)

    # -------------------------------------------------------------
    # 2. FOLHA: Rotas (As Rotas é o 2º!)
    # -------------------------------------------------------------
    ws_rotas = wb.create_sheet(title="Rotas")
    rotas_headers = [
        ("Rota", COLOR_REQ),
        ("Sequência", COLOR_REQ),
        ("Cliente", COLOR_REQ),
        ("Morada", COLOR_REQ),
        ("Código Postal", COLOR_REQ),
        ("Localidade", COLOR_REC),
        ("Contacto", COLOR_REC),
        ("Volume (m3)", COLOR_REC),
        ("Peso (kg)", COLOR_REC),
        ("Bultos/Caixas", COLOR_REC),
        ("Vendedor", COLOR_OPT),
        ("Observações", COLOR_OPT),
        ("Valor a Cobrar (COD)", COLOR_OPT),
        ("Janela_Inicio", COLOR_OPT),
        ("Janela_Fim", COLOR_OPT)
    ]
    rotas_data = [
        ["Rota Norte 1", 1, "Supermercado Central do Porto", "Rua de Santa Catarina 450", "4000-444", "Porto", "912345678", 1.2, 45.0, 3, "Manuel Pires", "Entregar nas traseiras / cais 2", 0.0, "08:30:00", "12:00:00"],
        ["Rota Norte 1", 2, "Restaurante Foz Velha", "Avenida do Brasil 120", "4150-151", "Porto", "934567890", 0.5, 15.0, 2, "Manuel Pires", "Recebe até às 12h30", 125.50, "09:00:00", "12:30:00"],
        ["Rota Norte 1", 3, "Farmácia Boavista", "Praça Mouzinho de Albuquerque 25", "4100-359", "Porto", "961122334", 0.3, 5.0, 1, "Ana Santos", "Falar com Dra. Maria", 0.0, "14:00:00", "18:00:00"],
        ["Rota Sul 2", 1, "Café & Bistrô Avenida", "Avenida da Liberdade 200", "1250-147", "Lisboa", "925566778", 0.8, 25.0, 2, "Rui Costa", "Pedir carimbo na fatura", 85.00, "09:00:00", "13:00:00"],
        ["Rota Sul 2", 2, "Hotel Baixa Chiado", "Rua Garrett 88", "1200-204", "Lisboa", "919988776", 2.5, 120.0, 6, "Rui Costa", "Entrada pelo elevador de serviço", 0.0, "10:00:00", "17:00:00"]
    ]
    apply_headers_and_data(ws_rotas, rotas_headers, rotas_data)

    # -------------------------------------------------------------
    # 3. FOLHA: Motoristas e Carros
    # -------------------------------------------------------------
    ws_drivers = wb.create_sheet(title="Motoristas e Carros")
    drivers_headers = [
        ("Motorista", COLOR_REQ),
        ("PIN/Password", COLOR_REQ),
        ("Viatura", COLOR_REC),
        ("Matrícula", COLOR_OPT),
        ("Telemóvel", COLOR_REC),
        ("Rota Atribuída", COLOR_OPT)
    ]
    drivers_data = [
        ["João Silva", "1111", "Mercedes Sprinter", "54-AB-12", "910000001", "Rota Norte 1"],
        ["Carlos Sousa", "2222", "Renault Master", "89-XY-34", "910000002", "Rota Sul 2"],
        ["António Ferreira", "3333", "Iveco Daily", "12-ZZ-99", "910000003", ""]
    ]
    apply_headers_and_data(ws_drivers, drivers_headers, drivers_data)

    # -------------------------------------------------------------
    # 4. FOLHA: Justificação entregas
    # -------------------------------------------------------------
    ws_reasons = wb.create_sheet(title="Justificação entregas")
    reasons_headers = [
        ("Motivos de Insucesso", COLOR_REQ),
        ("Código", COLOR_OPT),
        ("Descrição", COLOR_OPT)
    ]
    reasons_data = [
        ["Cliente Ausente / Fechado", "MOT-01", "Estabelecimento encerrado no horário previsto"],
        ["Carga Não Conforme / Danificada", "MOT-02", "Divergência de artigos ou avaria na mercadoria"],
        ["Cliente Recusou a Carga", "MOT-03", "Cliente não aceitou receber o pedido"],
        ["Morada Incorreta ou Incompleta", "MOT-04", "Morada não encontrada ou número de porta inexistente"],
        ["Sem Acesso / Rua em Obras", "MOT-05", "Viatura impossibilitada de aceder fisicamente ao local"],
        ["Fora do Horário de Descarga", "MOT-06", "Chegada após a janela limite de receção"],
        ["Falta de Pagamento (Cobrança)", "MOT-07", "Cliente sem meio de pagamento na entrega a cobrar"],
        ["Outro Motivo (ver notas)", "MOT-08", "Outra ocorrência especificada nas notas do motorista"]
    ]
    apply_headers_and_data(ws_reasons, reasons_headers, reasons_data)

    # -------------------------------------------------------------
    # 5. FOLHA: Instruções & Legenda
    # -------------------------------------------------------------
    ws_inst = wb.create_sheet(title="Instruções & Legenda")
    inst_headers = [
        ("Secção", COLOR_INSTRUCOES),
        ("Cor / Nível", COLOR_INSTRUCOES),
        ("Obrigatoriedade", COLOR_INSTRUCOES),
        ("Tipo de Dado", COLOR_INSTRUCOES),
        ("Exemplo", COLOR_INSTRUCOES),
        ("Regra de Negócio & Utilização", COLOR_INSTRUCOES)
    ]
    inst_data = [
        ["0. LEGENDA DE CORES", "Azul Escuro (#1E3A8A)", "OBRIGATÓRIO", "Visual", "Ex: Nome_Armazem, Rota, Cliente, Morada, CP, PIN", "Campos essenciais. A aplicação bloqueia ou falha se estiverem vazios."],
        ["0. LEGENDA DE CORES", "Azul Mais Claro (#2563EB)", "RECOMENDADO", "Visual", "Ex: Localidade, Contactos, Volume, Peso, Bultos, Viatura", "Aumenta a precisão e permite navegação telefónica e controlo de carga."],
        ["0. LEGENDA DE CORES", "Cinza Ardósia (#64748B)", "OPCIONAL", "Visual", "Ex: Latitude, Longitude, Observações, Janelas, Vendedor, COD", "Informação complementar que enriquece a experiência na estrada."],
        ["1. Armazém", "Nome_Armazem / Morada / CP", "OBRIGATÓRIO", "Texto / XXXX-XXX", "Armazém Central Porto / 4405-555", "Ponto de partida da distribuição e destino de descarga de sobras."],
        ["2. Rotas", "Rota / Sequência / Cliente", "OBRIGATÓRIO", "Texto / Número", "Rota Norte 1 / 1 / Supermercado Central", "Estrutura da rota que o motorista verá na lista do telemóvel."],
        ["2. Rotas", "Morada / Código Postal", "OBRIGATÓRIO", "Texto / XXXX-XXX", "Rua de Santa Catarina 450 / 4000-444", "Permite lançar a navegação GPS com 1 clique para o Google Maps."],
        ["2. Rotas", "Valor a Cobrar (COD)", "OPCIONAL", "Número Decimal", "125.50", "Alerta destacado em vermelho para o motorista cobrar no ato da entrega."],
        ["3. Motoristas", "Motorista / PIN", "OBRIGATÓRIO", "Texto / Código PIN", "João Silva / 1111", "Credenciais de acesso à WebApp PWA com vista exclusiva da sua rota."],
        ["4. Justificação", "Motivos de Insucesso", "OBRIGATÓRIO", "Texto", "Cliente Ausente / Fechado", "Opções que aparecem no telemóvel do motorista ao clicar 'Não Entregue'."]
    ]
    apply_headers_and_data(ws_inst, inst_headers, inst_data)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"[OK] Ficheiro Modelo AppGeoRoutePlan gerado em: {output_path}")

if __name__ == "__main__":
    # Save both in Ficheiros EXCEL and in motoristas_webapp/data
    root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    path1 = os.path.join(root_dir, "Ficheiros EXCEL", "AppGeoRoutePlan.xlsx")
    path2 = os.path.join(root_dir, "motoristas_webapp", "data", "AppGeoRoutePlan_Template.xlsx")
    
    build_app_georouteplan_template(path1)
    build_app_georouteplan_template(path2)
