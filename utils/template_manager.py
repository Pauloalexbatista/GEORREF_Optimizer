# -*- coding: utf-8 -*-
import os
import io
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_FAMILY = "Segoe UI"
FONT_HEADER = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
FONT_DATA = Font(name=FONT_FAMILY, size=10, color="1F2937")
FONT_DATA_BOLD = Font(name=FONT_FAMILY, size=10, bold=True, color="1F2937")

COLOR_REQ = "1E3A8A"        # Azul Escuro (OBRIGATÓRIO)
COLOR_REC = "2563EB"        # Azul Mais Claro (RECOMENDADO)
COLOR_OPT = "64748B"        # Cinza (OPCIONAL)
COLOR_MANIFESTO = "5B21B6"   # Púrpura
COLOR_INSTRUCOES = "1E293B" # Ardósia

FILL_ZEBRA = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER_SIDE = Side(border_style="thin", color="CBD5E1")
BORDER_CELL = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
BORDER_HEADER = Border(
    left=Side(border_style="thin", color="CBD5E1"),
    right=Side(border_style="thin", color="CBD5E1"),
    top=Side(border_style="medium", color="0F172A"),
    bottom=Side(border_style="medium", color="0F172A")
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_WRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_sheet_headers_tiered(ws, headers_spec):
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    for col_idx, (col_name, color_hex) in enumerate(headers_spec, start=1):
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        c = ws.cell(row=1, column=col_idx, value=col_name)
        c.font = FONT_HEADER
        c.fill = fill
        c.alignment = ALIGN_CENTER
        c.border = BORDER_HEADER

def autofit_columns(ws, max_cols=30, min_width=12):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col[0].column > max_cols:
            continue
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, min_width)

def format_data_rows(ws, start_row, end_row, num_cols, center_cols=(), right_cols=()):
    for r_idx in range(start_row, end_row + 1):
        ws.row_dimensions[r_idx].height = 20
        row_fill = FILL_ZEBRA if (r_idx % 2 == 0) else FILL_WHITE
        for c_idx in range(1, num_cols + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = FONT_DATA
            cell.fill = row_fill
            cell.border = BORDER_CELL
            if c_idx in center_cols:
                cell.alignment = ALIGN_CENTER
            elif c_idx in right_cols:
                cell.alignment = ALIGN_RIGHT
            else:
                cell.alignment = ALIGN_LEFT

# 1. ARMAZÉNS
def build_sheet_1_armazens(wb):
    ws = wb.active
    ws.title = "Armazéns"
    headers_spec = [
        ("Nome_Armazem", COLOR_REQ), ("Morada", COLOR_REQ), ("CP", COLOR_REQ),
        ("Localidade", COLOR_OPT), ("Latitude", COLOR_OPT), ("Longitude", COLOR_OPT),
        ("Hora_Abertura", COLOR_REC), ("Hora_Fecho", COLOR_REC), ("Tempo_Carga_Min", COLOR_REC),
        ("Contacto_Responsavel", COLOR_REC)
    ]
    apply_sheet_headers_tiered(ws, headers_spec)
    data = [
        ["Armazém Lisboa Central", "Avenida Severiano Falcão 16A", "2685-379", "Prior Velho", 38.7842, -9.1238, "06:00", "22:00", 30, "912345678"],
        ["Armazém Norte Maia", "Rua de Manuel Sousa Moreira Cruz 240", "4470-396", "Maia", 41.2291, -8.6620, "06:30", "21:30", 25, "923456789"]
    ]
    for r_idx, row_vals in enumerate(data, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws, 2, 1 + len(data), len(headers_spec), center_cols=(3, 7, 8, 9, 10), right_cols=(5, 6))
    autofit_columns(ws, len(headers_spec))

# 2. FROTA
def build_sheet_2_frota(wb):
    ws = wb.create_sheet("Frota")
    headers_spec = [
        ("Armazem", COLOR_REQ), ("Veiculo", COLOR_REQ), ("Capacidade_KG", COLOR_REQ),
        ("Capacidade_Vol", COLOR_REC), ("Velocidade_Media", COLOR_REC), ("Hora_Inicio_Turno", COLOR_REQ),
        ("Hora_Fim_Turno", COLOR_REQ), ("Custo_KM", COLOR_REC), ("Custo_Hora", COLOR_REC),
        ("Max_Entregas", COLOR_REC), ("Regras", COLOR_OPT), ("Motorista_Nome", COLOR_REC),
        ("Motorista_Telemovel", COLOR_REC)
    ]
    apply_sheet_headers_tiered(ws, headers_spec)
    data = [
        ["Armazém Lisboa Central", "Carrinha Pequena 01", 800, 6.0, 45, "08:00", "18:00", 0.65, 12.50, 30, "[PEQUENO]", "Manuel Silva", "910000001"],
        ["Armazém Lisboa Central", "Carrinha Longo Curso 02", 1200, 12.0, 55, "07:00", "19:30", 0.75, 14.00, 35, "[ALARGADO]", "António Santos", "910000002"]
    ]
    for r_idx, row_vals in enumerate(data, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws, 2, 1 + len(data), len(headers_spec), center_cols=(6, 7, 10, 11, 13), right_cols=(3, 4, 5, 8, 9))
    autofit_columns(ws, len(headers_spec))

# 3. ENTREGAS
def build_sheet_3_entregas(wb):
    ws = wb.create_sheet("Entregas")
    headers_spec = [
        ("Armazem", COLOR_REQ), ("Doc_ID", COLOR_REQ), ("Cliente", COLOR_REQ),
        ("Morada", COLOR_REQ), ("CP", COLOR_REQ), ("Localidade", COLOR_OPT),
        ("Latitude", COLOR_OPT), ("Longitude", COLOR_OPT), ("Telefone_Cliente", COLOR_REC),
        ("Peso_KG", COLOR_REC), ("Volume_M3", COLOR_REC), ("Janela1_Inicio", COLOR_REQ),
        ("Janela1_Fim", COLOR_REQ), ("Tempo_Descarga_Min", COLOR_REC), ("Regras", COLOR_OPT),
        ("Vendedor", COLOR_OPT), ("Valor_Cobrar", COLOR_OPT), ("Notas_Entrega", COLOR_OPT)
    ]
    apply_sheet_headers_tiered(ws, headers_spec)
    data = [
        ["Armazém Lisboa Central", "FT 2026/101", "Restaurante Alfama", "Rua de São Tomé 14", "1100-563", "Lisboa", 38.7138, -9.1301, "912345670", 80.0, 0.8, "09:00", "13:00", 15, "[PEQUENO]", "TNB", 0.0, "Entregar porta cais"],
        ["Armazém Lisboa Central", "FT 2026/102", "Hotel Avenida Palace", "Rua 1º de Dezembro 123", "1200-359", "Lisboa", 38.7150, -9.1415, "912345671", 320.0, 3.2, "08:30", "12:00", 25, "[ALARGADO]", "TNB", 0.0, "Recepção de mercadorias"]
    ]
    for r_idx, row_vals in enumerate(data, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws, 2, 1 + len(data), len(headers_spec), center_cols=(1, 2, 5, 9, 12, 13, 14, 15, 16), right_cols=(7, 8, 10, 11, 17))
    autofit_columns(ws, len(headers_spec))

# 4. REGRAS
def build_sheet_4_regras(wb):
    ws = wb.create_sheet("Regras")
    headers_spec = [
        ("Tag_Veiculo", COLOR_REQ), ("Permissao", COLOR_REQ), ("Tag_Entrega", COLOR_REQ), ("Descricao", COLOR_OPT)
    ]
    apply_sheet_headers_tiered(ws, headers_spec)
    data = [
        ["PEQUENO", "SIM", "PEQUENO", "Veículo pequeno acede a centros históricos e ruas estreitas"],
        ["ALARGADO", "NAO", "PEQUENO", "Veículo de longo curso proibido em ruas estreitas"]
    ]
    for r_idx, row_vals in enumerate(data, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws, 2, 1 + len(data), len(headers_spec), center_cols=(1, 2, 3))
    autofit_columns(ws, len(headers_spec))

# 5. ROTAS
def build_sheet_5_rotas(wb):
    ws = wb.create_sheet("Rotas")
    headers_spec = [
        ("ID_Original", COLOR_REQ), ("Cliente", COLOR_REQ), ("Morada", COLOR_REQ),
        ("Localidade", COLOR_REQ), ("CodPostal", COLOR_REQ), ("Rota", COLOR_REC),
        ("Ordem", COLOR_REC), ("Janela_Inicio", COLOR_REC), ("Janela_Fim", COLOR_REC),
        ("Peso", COLOR_REC), ("Volumes", COLOR_REC), ("Contacto", COLOR_REC),
        ("Vendedor", COLOR_OPT), ("Valor_Cobrar", COLOR_OPT), ("Observações", COLOR_OPT)
    ]
    apply_sheet_headers_tiered(ws, headers_spec)
    data = [
        ["D17932246", "Mário Pires", "Rua Manuel Antonio Gomes Lote 2 4K", "Lisboa", "1750-168", "Rota Norte 1", 1, "08:00", "10:00", 160.0, 4, "910000001", "TNB", 0.0, "Entregar porta 4K"],
        ["D17932606", "António Raposo Laura Furtado LDA", "Estrada Militar - Terminal Arnaud", "Camarate", "2680-183", "Rota Norte 1", 2, "09:00", "13:00", 286.0, 8, "910000002", "TNB", 0.0, "Cais traseiras"]
    ]
    for r_idx, row_vals in enumerate(data, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws, 2, 1 + len(data), len(headers_spec), center_cols=(1, 5, 6, 7, 8, 9, 11, 12, 13), right_cols=(10, 14))
    autofit_columns(ws, len(headers_spec))

# 6. MANIFESTOS
def build_sheet_6_manifestos(wb):
    ws = wb.create_sheet("Manifestos")
    headers_spec = [
        ("Armazem", COLOR_MANIFESTO), ("Veiculo", COLOR_MANIFESTO), ("Motorista", COLOR_MANIFESTO),
        ("Total_Paragens", COLOR_MANIFESTO), ("Volume_Total_M3", COLOR_MANIFESTO), ("Peso_Total_KG", COLOR_MANIFESTO),
        ("Taxa_Ocupacao_Peso", COLOR_MANIFESTO), ("Distancia_Total_KM", COLOR_MANIFESTO),
        ("Tempo_Total_Estimado", COLOR_MANIFESTO), ("Custo_Total_Estimado", COLOR_MANIFESTO),
        ("Lista_Documentos", COLOR_OPT)
    ]
    apply_sheet_headers_tiered(ws, headers_spec)
    data = [
        ["Armazém Lisboa Central", "Carrinha Pequena 01", "Manuel Silva", 2, 1.2, 130.0, "16.2%", 27.7, "02h 25m", 18.00, "FT 2026/101, FT 2026/104"]
    ]
    for r_idx, row_vals in enumerate(data, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws, 2, 1 + len(data), len(headers_spec), center_cols=(4, 7, 9), right_cols=(5, 6, 8, 10))
    autofit_columns(ws, len(headers_spec))

# 7. MOTORISTAS E CARROS
def build_sheet_7_motoristas(wb):
    ws = wb.create_sheet("Motoristas e Carros")
    headers_spec = [
        ("Motorista", COLOR_REQ), ("PIN/Password", COLOR_REQ), ("Viatura", COLOR_REC),
        ("Matrícula", COLOR_OPT), ("Telemóvel", COLOR_REC), ("Rota Atribuída", COLOR_OPT)
    ]
    apply_sheet_headers_tiered(ws, headers_spec)
    data = [
        ["João Silva", "1111", "Mercedes Sprinter", "54-AB-12", "910000001", "Rota Norte 1"],
        ["Carlos Sousa", "2222", "Renault Master", "89-XY-34", "910000002", "Rota Sul 2"],
        ["António Ferreira", "3333", "Iveco Daily", "12-ZZ-99", "910000003", ""]
    ]
    for r_idx, row_vals in enumerate(data, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws, 2, 1 + len(data), len(headers_spec), center_cols=(2, 4, 5, 6))
    autofit_columns(ws, len(headers_spec))

# 8. JUSTIFICAÇÃO ENTREGAS
def build_sheet_8_justificacoes(wb):
    ws = wb.create_sheet("Justificação entregas")
    headers_spec = [
        ("Motivo de Não Entrega", COLOR_REQ), ("Categoria / Ação", COLOR_REC)
    ]
    apply_sheet_headers_tiered(ws, headers_spec)
    data = [
        ["Cliente Ausente / Fechado", "Ausência"],
        ["Cliente Recusou a Carga", "Recusa"],
        ["Morada Não Encontrada / Errada", "Morada"],
        ["Mercadoria Danificada", "Avaria"],
        ["Falta de Tempo / Fora de Horas", "Operacional"],
        ["Sem Dinheiro para Cobrança", "Financeiro"]
    ]
    for r_idx, row_vals in enumerate(data, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws, 2, 1 + len(data), len(headers_spec), center_cols=(2,))
    autofit_columns(ws, len(headers_spec))

# 9. INSTRUÇÕES
def build_sheet_9_instrucoes(wb):
    ws = wb.create_sheet("Instruções")
    headers_spec = [
        ("Secção", COLOR_INSTRUCOES), ("Campo_ou_Regra", COLOR_INSTRUCOES),
        ("Obrigatorio", COLOR_INSTRUCOES), ("Formato_Aceite", COLOR_INSTRUCOES),
        ("Exemplo", COLOR_INSTRUCOES), ("Descricao_e_Recomendacoes", COLOR_INSTRUCOES)
    ]
    apply_sheet_headers_tiered(ws, headers_spec)
    instructions = [
        ["0. LEGENDA DE CORES", "Azul Escuro (#1E3A8A)", "OBRIGATÓRIO", "Visual", "Ex: Nome_Armazem, Morada, CP", "Campos essenciais para geocodificação e cálculo de rotas."],
        ["0. LEGENDA DE CORES", "Azul Mais Claro (#2563EB)", "RECOMENDADO", "Visual", "Ex: Janelas, Custos, Capacidade", "Aumenta a precisão de cálculo de tempos e indicadores."],
        ["0. LEGENDA DE CORES", "Cinza Ardósia (#64748B)", "OPCIONAL", "Visual", "Ex: Localidade, Notas, Valor a Cobrar", "Informação complementar."],
        ["1. Armazéns", "Nome_Armazem & CP", "SIM", "Texto / XXXX-XXX", "Armazém Central / 2685-379", "Identificador e morada do polo logístico."],
        ["2. Frota", "Veiculo & Capacidade_KG", "SIM", "Texto / Decimal", "Carrinha 01 / 800", "Viaturas e capacidades para otimização."],
        ["3. Entregas", "Cliente, Morada, CP, Janelas", "SIM", "Texto / HH:MM", "FT 2026/101 / 09:00 - 13:00", "Paragens a visitar."],
        ["4. Regras", "Tag_Veiculo & Tag_Entrega", "SIM", "Texto", "PEQUENO / SIM / PEQUENO", "Restrições de acesso."],
        ["5. Rotas", "Rotas Calculadas", "Leitura", "Ordem & Janelas", "Rota Norte 1 (#1)", "Sequência otimizada de entrega."],
        ["6. Manifestos", "Resumo de Carga", "Leitura", "Totais KM / Custo", "Total 27.7 km", "Folha de carga para motorista."],
        ["7. Motoristas e Carros", "Motorista & PIN", "SIM", "Texto / 4 dígitos", "João Silva / 1111", "Credenciais de acesso à WebApp móvel."],
        ["8. Justificação", "Motivo de Falha", "SIM", "Texto", "Cliente Ausente / Fechado", "Opções predefinidas de não entrega."]
    ]
    for r_idx, row_vals in enumerate(instructions, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws, 2, 1 + len(instructions), len(headers_spec), center_cols=(3, 4))
    autofit_columns(ws, len(headers_spec))

def create_unified_project_template() -> bytes:
    wb = openpyxl.Workbook()
    build_sheet_1_armazens(wb)
    build_sheet_2_frota(wb)
    build_sheet_3_entregas(wb)
    build_sheet_4_regras(wb)
    build_sheet_5_rotas(wb)
    build_sheet_6_manifestos(wb)
    build_sheet_7_motoristas(wb)
    build_sheet_8_justificacoes(wb)
    build_sheet_9_instrucoes(wb)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def save_master_template_file(output_path: str = "Ficheiros EXCEL/GeoRoutePlan.xlsx"):
    data = create_unified_project_template()
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "wb") as f:
        f.write(data)
    try:
        with open("Ficheiros EXCEL/AppGeoRoutePlan.xlsx", "wb") as f:
            f.write(data)
    except Exception:
        pass
    return output_path

if __name__ == "__main__":
    p = save_master_template_file()
    print(f"Generated COMPLETE 9-SHEET unified template at: {p}")
