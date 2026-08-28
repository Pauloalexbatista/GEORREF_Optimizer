# -*- coding: utf-8 -*-
"""
GEORREF Optimizer - Export Engine
Unified 7-Sheet Project Exporter matching GeoRoutePlan.xlsx standard
Universal Color Palette:
- Azul Escuro (#1E3A8A): OBRIGATÓRIO
- Azul Mais Claro (#2563EB): RECOMENDADO
- Cinza (#64748B): OPCIONAL
"""

import io
import math
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.template_manager import (
    FONT_HEADER, FONT_DATA, FONT_DATA_BOLD, FONT_MUTED,
    COLOR_REQ, COLOR_REC, COLOR_OPT, COLOR_MANIFESTO, COLOR_INSTRUCOES,
    FILL_ZEBRA, FILL_WHITE, BORDER_CELL, BORDER_HEADER,
    ALIGN_LEFT, ALIGN_CENTER, ALIGN_RIGHT, ALIGN_WRAP_LEFT,
    apply_sheet_headers_tiered, autofit_columns, format_data_rows, build_sheet_instrucoes
)

def safe_float(val, default=0.0) -> float:
    if val is None or pd.isna(val):
        return default
    try:
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return default
        return f
    except Exception:
        return default

def safe_int(val, default=0) -> int:
    if val is None or pd.isna(val):
        return default
    try:
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return default
        return int(f)
    except Exception:
        return default

def generate_full_project_excel(
    routes_df,
    deliveries_df=None,
    warehouses_df=None,
    fleet_config=None,
    optimization_params=None,
    rules_matrix=None,
    drivers_data=None,
    reasons_data=None
) -> bytes:
    wb = openpyxl.Workbook()
    
    # 1. Armazéns
    ws_wh = wb.active
    ws_wh.title = 'Armazéns'
    wh_headers_spec = [
        ('Nome_Armazem', COLOR_REQ),
        ('Morada', COLOR_REQ),
        ('CP', COLOR_REQ),
        ('Localidade', COLOR_OPT),
        ('Latitude', COLOR_OPT),
        ('Longitude', COLOR_OPT),
        ('Hora_Abertura', COLOR_REC),
        ('Hora_Fecho', COLOR_REC),
        ('Tempo_Carga_Min', COLOR_REC),
        ('Contacto_Responsavel', COLOR_REC)
    ]
    apply_sheet_headers_tiered(ws_wh, wh_headers_spec)
    
    wh_rows = []
    if warehouses_df is not None and not warehouses_df.empty:
        for _, r in warehouses_df.iterrows():
            wh_rows.append([
                str(r.get('Nome_Armazem', r.get('Nome', 'Armazém Central'))),
                str(r.get('Morada', '')),
                str(r.get('CP', r.get('Codigo_Postal', ''))),
                str(r.get('Localidade', r.get('Concelho', ''))),
                safe_float(r.get('Latitude')),
                safe_float(r.get('Longitude')),
                str(r.get('Hora_Abertura', '06:00:00')),
                str(r.get('Hora_Fecho', '22:00:00')),
                safe_int(r.get('Tempo_Carga_Min', 30)),
                str(r.get('Contacto_Responsavel', r.get('Contacto', '')))
            ])
    if not wh_rows:
        wh_rows = [['Armazém Central', 'Avenida Central 1', '1000-001', 'Lisboa', 38.7842, -9.1238, '06:00:00', '22:00:00', 30, '']]
    
    for r_idx, row_vals in enumerate(wh_rows, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws_wh.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws_wh, 2, 1 + len(wh_rows), len(wh_headers_spec), center_cols=(3, 7, 8, 9, 10), right_cols=(5, 6))
    autofit_columns(ws_wh, len(wh_headers_spec))
    
    # 2. Frota
    ws_fleet = wb.create_sheet('Frota')
    fleet_headers_spec = [
        ('Armazem', COLOR_REQ),
        ('Veiculo', COLOR_REQ),
        ('Capacidade_KG', COLOR_REQ),
        ('Capacidade_Vol', COLOR_REC),
        ('Velocidade_Media', COLOR_REC),
        ('Hora_Inicio_Turno', COLOR_REQ),
        ('Hora_Fim_Turno', COLOR_REQ),
        ('Custo_KM', COLOR_REC),
        ('Custo_Hora', COLOR_REC),
        ('Max_Entregas', COLOR_REC),
        ('Regras', COLOR_OPT),
        ('Motorista_Nome', COLOR_REC),
        ('Motorista_Telemovel', COLOR_REC)
    ]
    apply_sheet_headers_tiered(ws_fleet, fleet_headers_spec)
    
    fleet_rows = []
    if fleet_config is not None:
        if isinstance(fleet_config, pd.DataFrame):
            for _, r in fleet_config.iterrows():
                fleet_rows.append([
                    str(r.get('Armazem', 'Armazém Central')),
                    str(r.get('Veiculo', '')),
                    safe_float(r.get('Capacidade_KG', 1000.0)),
                    safe_float(r.get('Capacidade_Vol', r.get('Cap_Volume_m3', 10.0))),
                    safe_float(r.get('Velocidade_Media', 50.0)),
                    str(r.get('Hora_Inicio_Turno', r.get('Horario_Inicio', '08:00:00'))),
                    str(r.get('Hora_Fim_Turno', r.get('Horario_Fim', '18:00:00'))),
                    safe_float(r.get('Custo_KM', 0.65)),
                    safe_float(r.get('Custo_Hora', 12.50)),
                    safe_int(r.get('Max_Entregas', 30)),
                    str(r.get('Regras', '')),
                    str(r.get('Motorista_Nome', '')),
                    str(r.get('Motorista_Telemovel', ''))
                ])
        elif isinstance(fleet_config, dict):
            for v_name, v_data in fleet_config.items():
                if isinstance(v_data, dict):
                    fleet_rows.append([
                        str(v_data.get('armazem', 'Armazém Central')),
                        str(v_name),
                        safe_float(v_data.get('capacidade_kg', 1000.0)),
                        safe_float(v_data.get('capacidade_vol', 10.0)),
                        safe_float(v_data.get('velocidade_media', 50.0)),
                        str(v_data.get('horario_inicio', '08:00:00')),
                        str(v_data.get('horario_fim', '18:00:00')),
                        safe_float(v_data.get('custo_km', 0.65)),
                        safe_float(v_data.get('custo_hora', 12.50)),
                        safe_int(v_data.get('max_entregas', 30)),
                        str(v_data.get('regras', '')),
                        str(v_data.get('motorista_nome', '')),
                        str(v_data.get('motorista_telemovel', ''))
                    ])
                else:
                    fleet_rows.append([
                        str(getattr(v_data, 'armazem', 'Armazém Central')),
                        str(v_name),
                        safe_float(getattr(v_data, 'capacidade_kg', 1000.0)),
                        safe_float(getattr(v_data, 'capacidade_vol', 10.0)),
                        safe_float(getattr(v_data, 'velocidade_media', 50.0)),
                        str(getattr(v_data, 'horario_inicio', '08:00:00')),
                        str(getattr(v_data, 'horario_fim', '18:00:00')),
                        safe_float(getattr(v_data, 'custo_km', 0.65)),
                        safe_float(getattr(v_data, 'custo_hora', 12.50)),
                        safe_int(getattr(v_data, 'max_entregas', 30)),
                        str(getattr(v_data, 'regras', '')),
                        str(getattr(v_data, 'motorista_nome', '')),
                        str(getattr(v_data, 'motorista_telemovel', ''))
                    ])
    if not fleet_rows:
        fleet_rows = [['Armazém Central', 'Carrinha 01', 1000.0, 10.0, 50.0, '08:00:00', '18:00:00', 0.65, 12.50, 30, '', '', '']]
    
    for r_idx, row_vals in enumerate(fleet_rows, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws_fleet.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws_fleet, 2, 1 + len(fleet_rows), len(fleet_headers_spec), center_cols=(6, 7, 10, 11, 13), right_cols=(3, 4, 5, 8, 9))
    autofit_columns(ws_fleet, len(fleet_headers_spec))
    
    # 3. Entregas
    ws_entregas = wb.create_sheet('Entregas')
    deliv_headers_spec = [
        ('Armazem', COLOR_REQ),
        ('Doc_ID', COLOR_REQ),
        ('Cliente', COLOR_REQ),
        ('Morada', COLOR_REQ),
        ('CP', COLOR_REQ),
        ('Localidade', COLOR_OPT),
        ('Latitude', COLOR_OPT),
        ('Longitude', COLOR_OPT),
        ('Telefone_Cliente', COLOR_REC),
        ('Peso_KG', COLOR_REC),
        ('Volume_M3', COLOR_REC),
        ('Janela1_Inicio', COLOR_REQ),
        ('Janela1_Fim', COLOR_REQ),
        ('Janela2_Inicio', COLOR_OPT),
        ('Janela2_Fim', COLOR_OPT),
        ('Janela3_Inicio', COLOR_OPT),
        ('Janela3_Fim', COLOR_OPT),
        ('Tempo_Descarga_Min', COLOR_REC),
        ('Tipo_Operacao', COLOR_REC),
        ('Regras', COLOR_OPT),
        ('Notas_Motorista', COLOR_OPT),
        ('Prioridade', COLOR_REC),
        ('Vendedor', COLOR_OPT)
    ]
    apply_sheet_headers_tiered(ws_entregas, deliv_headers_spec)
    
    deliv_rows = []
    src_deliv = deliveries_df if (deliveries_df is not None and not deliveries_df.empty) else routes_df
    if src_deliv is not None and not src_deliv.empty:
        for _, r in src_deliv.iterrows():
            deliv_rows.append([
                str(r.get('Armazem', 'Armazém Central')),
                str(r.get('Doc_ID', r.get('Codigo_Cliente', r.get('Cliente', '')))),
                str(r.get('Cliente', r.get('Nome_Cliente', ''))),
                str(r.get('Morada', '')),
                str(r.get('CP', r.get('Codigo_Postal', ''))),
                str(r.get('Localidade', r.get('Concelho', ''))),
                safe_float(r.get('Latitude')),
                safe_float(r.get('Longitude')),
                str(r.get('Telefone_Cliente', r.get('Telefone', ''))),
                safe_float(r.get('Peso_KG', r.get('Peso', 0.0))),
                safe_float(r.get('Volume_M3', r.get('Volume_m3', 0.0))),
                str(r.get('Janela1_Inicio', r.get('Janela_Inicio', '08:00:00'))),
                str(r.get('Janela1_Fim', r.get('Janela_Fim', '18:00:00'))),
                str(r.get('Janela2_Inicio', '')),
                str(r.get('Janela2_Fim', '')),
                str(r.get('Janela3_Inicio', '')),
                str(r.get('Janela3_Fim', '')),
                safe_int(r.get('Tempo_Descarga_Min', 15)),
                str(r.get('Tipo_Operacao', 'Entrega')),
                str(r.get('Regras', '')),
                str(r.get('Notas_Motorista', r.get('Observacoes', ''))),
                str(r.get('Prioridade', 'Normal')),
                str(r.get('Vendedor', r.get('vendedor', '')))
            ])
    
    if deliv_rows:
        for r_idx, row_vals in enumerate(deliv_rows, start=2):
            for c_idx, val in enumerate(row_vals, start=1):
                ws_entregas.cell(row=r_idx, column=c_idx, value=val)
        format_data_rows(ws_entregas, 2, 1 + len(deliv_rows), len(deliv_headers_spec),
                         center_cols=(2, 5, 9, 12, 13, 14, 15, 16, 17, 18, 19, 20, 22),
                         right_cols=(7, 8, 10, 11))
    autofit_columns(ws_entregas, len(deliv_headers_spec))
    
    # 4. Regras
    ws_regras = wb.create_sheet('Regras')
    regras_headers_spec = [
        ('Tag_Veiculo', COLOR_REQ),
        ('Permissao', COLOR_REQ),
        ('Tag_Entrega', COLOR_REQ),
        ('Descricao', COLOR_OPT)
    ]
    apply_sheet_headers_tiered(ws_regras, regras_headers_spec)
    
    regras_data = []
    if rules_matrix and isinstance(rules_matrix, list):
        for item in rules_matrix:
            regras_data.append([
                str(item.get('Tag_Veiculo', item.get('tag_veiculo', ''))),
                str(item.get('Permissao', item.get('permissao', 'SIM'))).upper(),
                str(item.get('Tag_Entrega', item.get('tag_entrega', ''))),
                str(item.get('Descricao', item.get('descricao', '')))
            ])
    
    if regras_data:
        for r_idx, row_vals in enumerate(regras_data, start=2):
            for c_idx, val in enumerate(row_vals, start=1):
                ws_regras.cell(row=r_idx, column=c_idx, value=val)
        format_data_rows(ws_regras, 2, 1 + len(regras_data), len(regras_headers_spec), center_cols=(1, 2, 3))
    autofit_columns(ws_regras, len(regras_headers_spec))
    
    # 5. Rotas
    ws_rotas = wb.create_sheet('Rotas')
    rotas_headers_spec = [
        ('Armazem', COLOR_REQ),
        ('Veiculo', COLOR_REQ),
        ('Ordem_Paragem', COLOR_REQ),
        ('Doc_ID', COLOR_REC),
        ('Cliente', COLOR_REQ),
        ('Morada', COLOR_REQ),
        ('CP', COLOR_REQ),
        ('Localidade', COLOR_OPT),
        ('Telefone_Cliente', COLOR_REC),
        ('Janela_Horaria', COLOR_REC),
        ('Hora_Chegada_Prevista', COLOR_REC),
        ('Hora_Saida_Prevista', COLOR_REC),
        ('Distancia_KM', COLOR_REC),
        ('Distancia_Acumulada_KM', COLOR_REC),
        ('Tempo_Viagem_Min', COLOR_REC),
        ('Tempo_Espera_Min', COLOR_REC),
        ('Carga_Restante_KG', COLOR_REC),
        ('Carga_Restante_Vol', COLOR_REC),
        ('Status', COLOR_REC),
        ('Notas_Motorista', COLOR_OPT),
        ('Vendedor', COLOR_OPT)
    ]
    apply_sheet_headers_tiered(ws_rotas, rotas_headers_spec)
    
    rotas_rows = []
    manifest_map = {}
    
    default_wh_name = "Armazém Principal"
    if warehouses_df is not None and not (isinstance(warehouses_df, pd.DataFrame) and warehouses_df.empty):
        if isinstance(warehouses_df, pd.DataFrame) and not warehouses_df.empty:
            first_row = warehouses_df.iloc[0]
            default_wh_name = str(first_row.get("Nome_Armazem") or first_row.get("name") or first_row.get("Nome") or "Armazém Principal")
        elif isinstance(warehouses_df, list) and len(warehouses_df) > 0:
            first_item = warehouses_df[0]
            if isinstance(first_item, dict):
                default_wh_name = str(first_item.get("Nome_Armazem") or first_item.get("name") or "Armazém Principal")
    
    if routes_df is not None and not routes_df.empty:
        # Group by vehicle to sequence hours and distances
        route_groups = {}
        for idx, r in routes_df.iterrows():
            veh_name = str(r.get('Rota', r.get('Veiculo', 'Por Distribuir'))).strip()
            if veh_name not in route_groups:
                route_groups[veh_name] = []
            route_groups[veh_name].append((idx, r))
            
        for veh_name, items in route_groups.items():
            curr_minutes = 8 * 60  # 08:00
            cum_dist = 0.0
            cum_weight = 0.0
            
            for seq_idx, (orig_idx, r) in enumerate(items, start=1):
                wh_name = str(r.get('Armazem', default_wh_name)).strip()
                if not wh_name or wh_name in ['N/A', 'Armazém Central', 'Armazém Principal', 'None']:
                    wh_name = default_wh_name
                    
                ordem = safe_int(r.get('Ordem', seq_idx))
                doc_id = str(r.get('Doc_ID', r.get('Codigo_Cliente', f"DOC_{seq_idx}")))
                cliente = str(r.get('Cliente', r.get('Nome_Cliente', '')))
                morada = str(r.get('Morada', ''))
                cp = str(r.get('CP', r.get('Codigo_Postal', '')))
                loc = str(r.get('Localidade', r.get('Concelho', '')))
                tel = str(r.get('Telefone_Cliente', r.get('Telefone', '')))
                janela = str(r.get('Janela_Horaria', '08:00 - 18:00'))
                
                # Dynamic realistic schedule calculation
                leg_km = safe_float(r.get('KM_Anterior', r.get('Distancia_KM', 8.5)))
                if leg_km <= 0:
                    leg_km = 8.5
                cum_dist += leg_km
                
                t_viagem = safe_int(r.get('Tempo_Viagem_Min', 12))
                if t_viagem <= 0:
                    t_viagem = max(5, int(leg_km * 1.5))
                curr_minutes += t_viagem
                
                chegada_str = r.get('Chegada') or r.get('Hora_Chegada_Prevista')
                if not chegada_str or str(chegada_str).strip() in ['00:00', '08:00:00']:
                    arr_h = int(curr_minutes // 60)
                    arr_m = int(curr_minutes % 60)
                    chegada = f"{arr_h:02d}:{arr_m:02d}:00"
                else:
                    chegada = str(chegada_str)
                    
                service_min = 15
                curr_minutes += service_min
                
                saida_str = r.get('Saida') or r.get('Hora_Saida_Prevista')
                if not saida_str or str(saida_str).strip() in ['00:00', '08:15:00']:
                    dep_h = int(curr_minutes // 60)
                    dep_m = int(curr_minutes % 60)
                    saida = f"{dep_h:02d}:{dep_m:02d}:00"
                else:
                    saida = str(saida_str)
                    
                dist_km = round(leg_km, 2)
                dist_acum = round(cum_dist, 2)
                t_espera = safe_int(r.get('Tempo_Espera', r.get('Tempo_Espera_Min', 0)))
                
                p_kg = safe_float(r.get('Peso_KG', r.get('Peso', 45.0)))
                cum_weight += p_kg
                carga_kg = round(cum_weight, 1)
                carga_vol = safe_float(r.get('Volume_m3', r.get('Volume_M3', 0.15)))
                status = str(r.get('Status', 'Pendente'))
                
                driver_notes = str(r.get('Notas_Motorista', r.get('Observacoes', r.get('observacoes', ''))))
                vendedor_name = str(r.get('Vendedor', r.get('vendedor', '')))
                
                rotas_rows.append([
                    wh_name, veh_name, ordem, doc_id, cliente, morada, cp, loc, tel, janela,
                    chegada, saida, dist_km, dist_acum, t_viagem, t_espera, carga_kg, carga_vol, status,
                    driver_notes, vendedor_name
                ])
            
            if veh_name not in manifest_map:
                manifest_map[veh_name] = {
                    'armazem': wh_name,
                    'veiculo': veh_name,
                    'motorista': '',
                    'total_clientes': 0,
                    'total_vol': 0.0,
                    'total_peso': 0.0,
                    'total_km': 0.0,
                    'docs': []
                }
            manifest_map[veh_name]['total_clientes'] += 1
            manifest_map[veh_name]['total_peso'] += safe_float(r.get('Peso_KG', 0.0))
            manifest_map[veh_name]['total_vol'] += safe_float(r.get('Volume_M3', 0.0))
            manifest_map[veh_name]['total_km'] = max(manifest_map[veh_name]['total_km'], dist_acum)
            if doc_id and doc_id not in ['PARTIDA', 'RETORNO']:
                manifest_map[veh_name]['docs'].append(doc_id)
    
    if rotas_rows:
        for r_idx, row_vals in enumerate(rotas_rows, start=2):
            for c_idx, val in enumerate(row_vals, start=1):
                ws_rotas.cell(row=r_idx, column=c_idx, value=val)
        format_data_rows(ws_rotas, 2, 1 + len(rotas_rows), len(rotas_headers_spec),
                         center_cols=(3, 4, 7, 9, 10, 11, 12, 19),
                         right_cols=(13, 14, 15, 16, 17, 18))
    autofit_columns(ws_rotas, len(rotas_headers_spec))
    
    # 6. Manifestos
    ws_manifest = wb.create_sheet('Manifestos')
    manifest_headers_spec = [
        ('Armazem', COLOR_MANIFESTO),
        ('Veiculo', COLOR_MANIFESTO),
        ('Motorista', COLOR_MANIFESTO),
        ('Total_Clientes', COLOR_MANIFESTO),
        ('Total_Volumes', COLOR_MANIFESTO),
        ('Total_Peso_KG', COLOR_MANIFESTO),
        ('%_Ocupacao_Capacidade', COLOR_MANIFESTO),
        ('Total_KM_Estimados', COLOR_MANIFESTO),
        ('Total_Tempo_Turno', COLOR_MANIFESTO),
        ('Custo_Estimado_Total', COLOR_MANIFESTO),
        ('Lista_Documentos_Transporte', COLOR_MANIFESTO)
    ]
    apply_sheet_headers_tiered(ws_manifest, manifest_headers_spec)
    
    manifest_rows = []
    for v_name, m_info in manifest_map.items():
        doc_summary = ', '.join(m_info['docs'][:15]) + ('...' if len(m_info['docs']) > 15 else '')
        manifest_rows.append([
            m_info['armazem'],
            v_name,
            m_info.get('motorista', ''),
            m_info['total_clientes'],
            round(m_info['total_vol'], 2),
            round(m_info['total_peso'], 1),
            'N/D',
            round(m_info['total_km'], 1),
            '--:--',
            0.0,
            doc_summary
        ])
    
    if manifest_rows:
        for r_idx, row_vals in enumerate(manifest_rows, start=2):
            for c_idx, val in enumerate(row_vals, start=1):
                ws_manifest.cell(row=r_idx, column=c_idx, value=val)
        format_data_rows(ws_manifest, 2, 1 + len(manifest_rows), len(manifest_headers_spec),
                         center_cols=(4, 7, 9), right_cols=(5, 6, 8), currency_cols=(10,))
    autofit_columns(ws_manifest, len(manifest_headers_spec))
    
        # 7. Motoristas e Carros
    ws_drivers = wb.create_sheet('Motoristas e Carros')
    drivers_headers_spec = [
        ('Motorista', COLOR_REQ),
        ('PIN/Password', COLOR_REQ),
        ('Viatura', COLOR_REC),
        ('Matrícula', COLOR_OPT),
        ('Telemóvel', COLOR_REC),
        ('Rota Atribuída', COLOR_OPT)
    ]
    apply_sheet_headers_tiered(ws_drivers, drivers_headers_spec)
    
    drivers_rows = []
    if drivers_data:
        if isinstance(drivers_data, list):
            for d in drivers_data:
                if isinstance(d, dict):
                    drivers_rows.append([
                        str(d.get('name', d.get('Motorista', ''))),
                        str(d.get('pin', d.get('PIN/Password', d.get('PIN', '1234')))),
                        str(d.get('vehicle', d.get('Viatura', ''))),
                        str(d.get('matricula', d.get('Matrícula', ''))),
                        str(d.get('phone', d.get('Telemóvel', d.get('telefone', '')))),
                        str(d.get('route', d.get('Rota Atribuída', '')))
                    ])
        elif isinstance(drivers_data, pd.DataFrame):
            for _, d in drivers_data.iterrows():
                drivers_rows.append([
                    str(d.get('Motorista', d.get('name', ''))),
                    str(d.get('PIN/Password', d.get('pin', '1234'))),
                    str(d.get('Viatura', d.get('vehicle', ''))),
                    str(d.get('Matrícula', d.get('matricula', ''))),
                    str(d.get('Telemóvel', d.get('phone', ''))),
                    str(d.get('Rota Atribuída', d.get('route', '')))
                ])
                
    if not drivers_rows and fleet_config:
        # Fallback to fleet driver names if no explicit driver list
        if isinstance(fleet_config, dict):
            for v_name, v_data in fleet_config.items():
                d_name = v_data.get('motorista_nome', '') if isinstance(v_data, dict) else getattr(v_data, 'motorista_nome', '')
                d_tel = v_data.get('motorista_telemovel', '') if isinstance(v_data, dict) else getattr(v_data, 'motorista_telemovel', '')
                if d_name:
                    drivers_rows.append([d_name, '1111', v_name, '', d_tel, ''])
                    
    if not drivers_rows:
        drivers_rows = [
            ['Motorista 1', '1111', 'Viatura 1', '', '910000001', 'Rota 1']
        ]
        
    for r_idx, row_vals in enumerate(drivers_rows, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws_drivers.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws_drivers, 2, 1 + len(drivers_rows), len(drivers_headers_spec), center_cols=(2, 4, 5, 6))
    autofit_columns(ws_drivers, len(drivers_headers_spec))
    
    # 8. Justificação entregas
    ws_reasons = wb.create_sheet('Justificação entregas')
    reasons_headers_spec = [
        ('Motivo de Não Entrega', COLOR_REQ),
        ('Categoria / Ação', COLOR_REC)
    ]
    apply_sheet_headers_tiered(ws_reasons, reasons_headers_spec)
    
    reasons_rows = []
    if reasons_data:
        if isinstance(reasons_data, list):
            for r in reasons_data:
                if isinstance(r, dict):
                    reasons_rows.append([
                        str(r.get('reason', r.get('Motivo de Não Entrega', r.get('motivo', '')))),
                        str(r.get('category', r.get('Categoria / Ação', r.get('categoria', 'Geral'))))
                    ])
        elif isinstance(reasons_data, pd.DataFrame):
            for _, r in reasons_data.iterrows():
                reasons_rows.append([
                    str(r.get('Motivo de Não Entrega', r.get('reason', ''))),
                    str(r.get('Categoria / Ação', r.get('category', 'Geral')))
                ])
                
    if not reasons_rows:
        reasons_rows = [
            ['Cliente Ausente / Fechado', 'Ausência'],
            ['Cliente Recusou a Carga', 'Recusa'],
            ['Morada Não Encontrada / Errada', 'Morada'],
            ['Mercadoria Danificada', 'Avaria'],
            ['Falta de Tempo / Fora de Horas', 'Operacional'],
            ['Sem Dinheiro para Cobrança', 'Financeiro']
        ]
        
    for r_idx, row_vals in enumerate(reasons_rows, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws_reasons.cell(row=r_idx, column=c_idx, value=val)
    format_data_rows(ws_reasons, 2, 1 + len(reasons_rows), len(reasons_headers_spec), center_cols=(2,))
    autofit_columns(ws_reasons, len(reasons_headers_spec))

    # 9. Instruções
    build_sheet_instrucoes(wb)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_route_excel(routes_df):
    return generate_full_project_excel(routes_df=routes_df)
